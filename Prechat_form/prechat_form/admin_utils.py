import csv
import io
import json
from datetime import datetime
from django.http import HttpResponse
from django.contrib import admin
from django.utils.html import format_html
from django.urls import path, reverse
from django.shortcuts import render, redirect
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.decorators import method_decorator
from django.db.models import Count, Q
from django.utils import timezone
from datetime import timedelta

# Excel/XLSX support
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# PDF support
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False


class DataExportMixin:
    """Mixin to add data export functionality to admin classes"""
    
    def export_csv(self, request, queryset):
        """Export selected items to CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{self.model._meta.verbose_name_plural}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        
        # Get field names
        field_names = [field.name for field in self.model._meta.fields]
        writer.writerow(field_names)
        
        # Write data
        for obj in queryset:
            row = []
            for field_name in field_names:
                value = getattr(obj, field_name)
                if hasattr(value, 'strftime'):  # Handle datetime fields
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif value is None:
                    value = ''
                row.append(str(value))
            writer.writerow(row)
        
        return response
    export_csv.short_description = "Export selected to CSV"
    
    def export_excel(self, request, queryset):
        """Export selected items to Excel"""
        if not EXCEL_AVAILABLE:
            self.message_user(request, "Excel export not available. Install openpyxl.", level='error')
            return
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.model._meta.verbose_name_plural
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Get field names and write headers
        field_names = [field.verbose_name.title() for field in self.model._meta.fields]
        for col, field_name in enumerate(field_names, 1):
            cell = ws.cell(row=1, column=col, value=field_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row_num, obj in enumerate(queryset, 2):
            for col, field in enumerate(self.model._meta.fields, 1):
                value = getattr(obj, field.name)
                if hasattr(value, 'strftime'):  # Handle datetime fields
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif value is None:
                    value = ''
                ws.cell(row=row_num, column=col, value=str(value))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.model._meta.verbose_name_plural}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
    export_excel.short_description = "Export selected to Excel"
    
    def export_pdf(self, request, queryset):
        """Export selected items to PDF"""
        if not PDF_AVAILABLE:
            self.message_user(request, "PDF export not available. Install reportlab.", level='error')
            return
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{self.model._meta.verbose_name_plural}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        # Create PDF document
        doc = SimpleDocTemplate(response, pagesize=A4)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Title
        title = Paragraph(f"{self.model._meta.verbose_name_plural} Report", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Prepare table data
        field_names = [field.verbose_name.title() for field in self.model._meta.fields]
        data = [field_names]
        
        for obj in queryset:
            row = []
            for field in self.model._meta.fields:
                value = getattr(obj, field.name)
                if hasattr(value, 'strftime'):  # Handle datetime fields
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif value is None:
                    value = ''
                # Truncate long text for PDF
                text = str(value)
                if len(text) > 50:
                    text = text[:47] + "..."
                row.append(text)
            data.append(row)
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        return response
    export_pdf.short_description = "Export selected to PDF"
    
    def export_json(self, request, queryset):
        """Export selected items to JSON"""
        data = []
        for obj in queryset:
            item = {}
            for field in self.model._meta.fields:
                value = getattr(obj, field.name)
                if hasattr(value, 'strftime'):  # Handle datetime fields
                    value = value.isoformat()
                elif hasattr(value, '__str__'):
                    value = str(value)
                item[field.name] = value
            data.append(item)
        
        response = HttpResponse(
            json.dumps(data, indent=2, ensure_ascii=False),
            content_type='application/json'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.model._meta.verbose_name_plural}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json"'
        
        return response
    export_json.short_description = "Export selected to JSON"


def get_admin_stats():
    """Get statistics for admin dashboard"""
    from .models import UserContact, ChatSession, FormSubmission, IntegrationLog
    
    # Date ranges
    today = timezone.now().date()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)
    
    stats = {
        'total_users': UserContact.objects.count(),
        'total_sessions': ChatSession.objects.count(),
        'total_submissions': FormSubmission.objects.count(),
        'users_today': UserContact.objects.filter(created_at__date=today).count(),
        'users_this_week': UserContact.objects.filter(created_at__date__gte=week_ago).count(),
        'users_this_month': UserContact.objects.filter(created_at__date__gte=month_ago).count(),
        'active_sessions': ChatSession.objects.filter(status='active').count(),
        'successful_submissions': FormSubmission.objects.filter(status='success').count(),
        'failed_submissions': FormSubmission.objects.filter(status__in=['validation_error', 'server_error']).count(),
        'recent_errors': IntegrationLog.objects.filter(level='error', created_at__date__gte=week_ago).count(),
    }
    
    return stats


@staff_member_required
def admin_dashboard(request):
    """Custom admin dashboard with statistics"""
    stats = get_admin_stats()
    
    context = {
        'title': 'PreChat Form Dashboard',
        'stats': stats,
        'has_permission': True,
    }
    
    return render(request, 'admin/dashboard.html', context)


@staff_member_required
def bulk_export_view(request):
    """Bulk export view for all data"""
    from .models import UserContact, ChatSession, FormSubmission, IntegrationLog
    
    if request.method == 'POST':
        export_type = request.POST.get('export_type')
        models_to_export = request.POST.getlist('models')
        
        if export_type == 'csv':
            return export_all_csv(models_to_export)
        elif export_type == 'excel':
            return export_all_excel(models_to_export)
        elif export_type == 'pdf':
            return export_all_pdf(models_to_export)
        elif export_type == 'json':
            return export_all_json(models_to_export)
    
    context = {
        'title': 'Bulk Data Export',
        'models': [
            ('UserContact', 'User Contacts'),
            ('ChatSession', 'Chat Sessions'),
            ('FormSubmission', 'Form Submissions'),
            ('IntegrationLog', 'Integration Logs'),
        ],
        'has_permission': True,
    }
    
    return render(request, 'admin/bulk_export.html', context)


def export_all_csv(models_to_export):
    """Export all selected models to a single CSV file"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="prechat_data_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    
    from .models import UserContact, ChatSession, FormSubmission, IntegrationLog
    model_map = {
        'UserContact': UserContact,
        'ChatSession': ChatSession,
        'FormSubmission': FormSubmission,
        'IntegrationLog': IntegrationLog,
    }
    
    for model_name in models_to_export:
        if model_name in model_map:
            model = model_map[model_name]
            
            # Write model header
            writer.writerow([f"=== {model._meta.verbose_name_plural.upper()} ==="])
            
            # Write field headers
            field_names = [field.name for field in model._meta.fields]
            writer.writerow(field_names)
            
            # Write data
            for obj in model.objects.all():
                row = []
                for field_name in field_names:
                    value = getattr(obj, field_name)
                    if hasattr(value, 'strftime'):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    elif value is None:
                        value = ''
                    row.append(str(value))
                writer.writerow(row)
            
            # Add separator
            writer.writerow([])
    
    return response


def export_all_excel(models_to_export):
    """Export all selected models to Excel with multiple sheets"""
    if not EXCEL_AVAILABLE:
        return HttpResponse("Excel export not available", status=400)
    
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    from .models import UserContact, ChatSession, FormSubmission, IntegrationLog
    model_map = {
        'UserContact': UserContact,
        'ChatSession': ChatSession,
        'FormSubmission': FormSubmission,
        'IntegrationLog': IntegrationLog,
    }
    
    for model_name in models_to_export:
        if model_name in model_map:
            model = model_map[model_name]
            ws = wb.create_sheet(title=model._meta.verbose_name_plural[:30])
            
            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Write headers
            field_names = [field.verbose_name.title() for field in model._meta.fields]
            for col, field_name in enumerate(field_names, 1):
                cell = ws.cell(row=1, column=col, value=field_name)
                cell.font = header_font
                cell.fill = header_fill
            
            # Write data
            for row_num, obj in enumerate(model.objects.all(), 2):
                for col, field in enumerate(model._meta.fields, 1):
                    value = getattr(obj, field.name)
                    if hasattr(value, 'strftime'):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    elif value is None:
                        value = ''
                    ws.cell(row=row_num, column=col, value=str(value))
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="prechat_data_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


def export_all_json(models_to_export):
    """Export all selected models to JSON"""
    data = {}
    
    from .models import UserContact, ChatSession, FormSubmission, IntegrationLog
    model_map = {
        'UserContact': UserContact,
        'ChatSession': ChatSession,
        'FormSubmission': FormSubmission,
        'IntegrationLog': IntegrationLog,
    }
    
    for model_name in models_to_export:
        if model_name in model_map:
            model = model_map[model_name]
            model_data = []
            
            for obj in model.objects.all():
                item = {}
                for field in model._meta.fields:
                    value = getattr(obj, field.name)
                    if hasattr(value, 'strftime'):
                        value = value.isoformat()
                    elif hasattr(value, '__str__'):
                        value = str(value)
                    item[field.name] = value
                model_data.append(item)
            
            data[model._meta.verbose_name_plural] = model_data
    
    response = HttpResponse(
        json.dumps(data, indent=2, ensure_ascii=False),
        content_type='application/json'
    )
    response['Content-Disposition'] = f'attachment; filename="prechat_data_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json"'
    
    return response


def export_all_pdf(models_to_export):
    """Export all selected models to PDF"""
    if not PDF_AVAILABLE:
        return HttpResponse("PDF export not available", status=400)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="prechat_data_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1
    )
    
    # Main title
    title = Paragraph("PreChat Form Data Export", title_style)
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    from .models import UserContact, ChatSession, FormSubmission, IntegrationLog
    model_map = {
        'UserContact': UserContact,
        'ChatSession': ChatSession,
        'FormSubmission': FormSubmission,
        'IntegrationLog': IntegrationLog,
    }
    
    for model_name in models_to_export:
        if model_name in model_map:
            model = model_map[model_name]
            
            # Model title
            model_title = Paragraph(f"{model._meta.verbose_name_plural}", styles['Heading2'])
            elements.append(model_title)
            elements.append(Spacer(1, 12))
            
            # Prepare table data (limit for PDF readability)
            field_names = [field.verbose_name.title() for field in model._meta.fields[:6]]  # Limit columns
            data = [field_names]
            
            for obj in model.objects.all()[:50]:  # Limit rows
                row = []
                for field in model._meta.fields[:6]:
                    value = getattr(obj, field.name)
                    if hasattr(value, 'strftime'):
                        value = value.strftime('%Y-%m-%d')
                    elif value is None:
                        value = ''
                    text = str(value)
                    if len(text) > 30:
                        text = text[:27] + "..."
                    row.append(text)
                data.append(row)
            
            # Create table
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
            elements.append(Spacer(1, 20))
    
    doc.build(elements)
    return response


@staff_member_required
def export_data_view(request, model_name, format_type, object_id=None):
    """Export data view for individual records or entire models"""
    from .models import UserContact, ChatSession, FormSubmission, IntegrationLog
    
    model_map = {
        'usercontact': UserContact,
        'chatsession': ChatSession,
        'formsubmission': FormSubmission,
        'integrationlog': IntegrationLog,
    }
    
    model_name_lower = model_name.lower()
    if model_name_lower not in model_map:
        return HttpResponse("Invalid model", status=400)
    
    model = model_map[model_name_lower]
    
    # Get queryset
    if object_id:
        try:
            queryset = model.objects.filter(id=object_id)
            if not queryset.exists():
                return HttpResponse("Object not found", status=404)
        except (ValueError, model.DoesNotExist):
            return HttpResponse("Invalid object ID", status=400)
    else:
        queryset = model.objects.all()
    
    # Export based on format
    if format_type == 'csv':
        return _export_queryset_csv(queryset, model)
    elif format_type == 'excel':
        return _export_queryset_excel(queryset, model)
    elif format_type == 'pdf':
        return _export_queryset_pdf(queryset, model)
    elif format_type == 'json':
        return _export_queryset_json(queryset, model)
    else:
        return HttpResponse("Invalid format", status=400)


def _export_queryset_csv(queryset, model):
    """Export queryset to CSV"""
    response = HttpResponse(content_type='text/csv')
    filename = f"{model._meta.verbose_name_plural}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Write headers
    field_names = [field.name for field in model._meta.fields]
    writer.writerow(field_names)
    
    # Write data
    for obj in queryset:
        row = []
        for field_name in field_names:
            value = getattr(obj, field_name)
            if hasattr(value, 'strftime'):
                value = value.strftime('%Y-%m-%d %H:%M:%S')
            elif value is None:
                value = ''
            row.append(str(value))
        writer.writerow(row)
    
    return response


def _export_queryset_excel(queryset, model):
    """Export queryset to Excel"""
    if not EXCEL_AVAILABLE:
        return HttpResponse("Excel export not available", status=400)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = model._meta.verbose_name_plural[:30]
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Write headers
    field_names = [field.verbose_name.title() for field in model._meta.fields]
    for col, field_name in enumerate(field_names, 1):
        cell = ws.cell(row=1, column=col, value=field_name)
        cell.font = header_font
        cell.fill = header_fill
    
    # Write data
    for row_num, obj in enumerate(queryset, 2):
        for col, field in enumerate(model._meta.fields, 1):
            value = getattr(obj, field.name)
            if hasattr(value, 'strftime'):
                value = value.strftime('%Y-%m-%d %H:%M:%S')
            elif value is None:
                value = ''
            ws.cell(row=row_num, column=col, value=str(value))
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{model._meta.verbose_name_plural}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


def _export_queryset_json(queryset, model):
    """Export queryset to JSON"""
    data = []
    
    for obj in queryset:
        item = {}
        for field in model._meta.fields:
            value = getattr(obj, field.name)
            if hasattr(value, 'strftime'):
                value = value.isoformat()
            elif hasattr(value, '__str__'):
                value = str(value)
            item[field.name] = value
        data.append(item)
    
    response = HttpResponse(
        json.dumps(data, indent=2, ensure_ascii=False),
        content_type='application/json'
    )
    filename = f"{model._meta.verbose_name_plural}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def _export_queryset_pdf(queryset, model):
    """Export queryset to PDF"""
    if not PDF_AVAILABLE:
        return HttpResponse("PDF export not available", status=400)
    
    response = HttpResponse(content_type='application/pdf')
    filename = f"{model._meta.verbose_name_plural}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=20,
        alignment=1
    )
    
    # Title
    title = Paragraph(f"{model._meta.verbose_name_plural} Export", title_style)
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    # Prepare table data (limit for PDF readability)
    field_names = [field.verbose_name.title() for field in model._meta.fields[:6]]
    data = [field_names]
    
    for obj in queryset[:100]:  # Limit rows for PDF
        row = []
        for field in model._meta.fields[:6]:
            value = getattr(obj, field.name)
            if hasattr(value, 'strftime'):
                value = value.strftime('%Y-%m-%d')
            elif value is None:
                value = ''
            text = str(value)
            if len(text) > 30:
                text = text[:27] + "..."
            row.append(text)
        data.append(row)
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    return response
